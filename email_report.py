"""
Email the latest comparison_report_*.xlsx to a fixed recipient.

Designed to be run on a schedule (Windows Task Scheduler) while main.py
is still writing to the report. We COPY the file - never rename or move
it - so the running automation keeps saving to the same path without
interference.

Idempotent: a small state file records the last (mtime, size) we sent;
runs that find no change since the previous send skip silently.
"""

import datetime
import glob
import logging
import os
import shutil
import smtplib
import sys
import tempfile
import time
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl

# ── Config ────────────────────────────────────────────────────────────────
SENDER       = "dirccmospi@gmail.com"
APP_PASSWORD = "PASTE_APP_PASSWORD_HERE"      # fill in locally; never commit
RECIPIENTS   = ["sarthaksri@gmail.com"]

SCRIPT_DIR   = Path(__file__).resolve().parent
REPORT_GLOB  = str(SCRIPT_DIR / "reports" / "comparison_report_*.xlsx")
STATE_FILE   = SCRIPT_DIR / ".email_report_state.txt"

SMTP_HOST    = "smtp.gmail.com"
SMTP_PORT    = 587


# ── Helpers ───────────────────────────────────────────────────────────────

def find_latest_report() -> Path | None:
    files = glob.glob(REPORT_GLOB)
    if not files:
        return None
    return Path(max(files, key=os.path.getmtime))


def has_changed(path: Path) -> bool:
    """True if (mtime_ns, size) differs from the last successful send."""
    st  = path.stat()
    sig = f"{st.st_mtime_ns}:{st.st_size}"
    last = STATE_FILE.read_text().strip() if STATE_FILE.exists() else ""
    if sig == last:
        return False
    STATE_FILE.write_text(sig)
    return True


def summarise(path: Path) -> str:
    """Short stats block included in the email body."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return f"(Could not summarise: {exc})"

    pass_n = fail_n = err_n = total = 0
    last_table = ""
    if "Summary" in wb.sheetnames:
        for row in wb["Summary"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            total += 1
            status = str(row[6] or "").upper() if len(row) > 6 else ""
            if   status == "PASS":  pass_n += 1
            elif status == "FAIL":  fail_n += 1
            elif status == "ERROR": err_n  += 1
            if len(row) > 2 and row[2]:
                last_table = str(row[2])[:80]

    mismatches = 0
    if "Mismatches" in wb.sheetnames:
        mismatches = max(0, wb["Mismatches"].max_row - 1)

    completed = []
    if "Completed" in wb.sheetnames:
        for row in wb["Completed"].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                completed.append(str(row[0]))

    wb.close()

    lines = [
        f"Tables processed:   {total}  (PASS {pass_n} | FAIL {fail_n} | ERROR {err_n})",
        f"Mismatch rows:      {mismatches}",
        f"Datasets completed: {len(completed)}" + (f" - {', '.join(completed)}" if completed else ""),
    ]
    if last_table:
        lines.append(f"Latest table:       {last_table}")
    return "\n".join(lines)


def send_email(body: str, attachment_path: Path, attachment_name: str) -> None:
    msg = MIMEMultipart()
    msg["From"]    = SENDER
    msg["To"]      = ", ".join(RECIPIENTS)
    msg["Subject"] = f"MoSPI Catalogue Automation - progress {datetime.datetime.now():%Y-%m-%d %H:%M}"
    msg.attach(MIMEText(body, "plain"))

    with open(attachment_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{attachment_name}"')
    msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
        s.starttls()
        s.login(SENDER, APP_PASSWORD)
        s.sendmail(SENDER, RECIPIENTS, msg.as_string())


def main() -> int:
    logging.basicConfig(
        format="%(asctime)s %(levelname)s %(message)s",
        level=logging.INFO,
    )

    src = find_latest_report()
    if src is None:
        logging.info("No report file matching %s - nothing to send.", REPORT_GLOB)
        return 0

    if not has_changed(src):
        logging.info("Report unchanged since last send (%s) - skipping.", src.name)
        return 0

    ts        = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    snap_name = f"{src.stem}_snapshot_{ts}.xlsx"

    # COPY to a temp dir so main.py keeps writing to the original. The
    # one race condition is main.py being mid-save when we copy - retry
    # once after a short pause to dodge that 1-2s window.
    with tempfile.TemporaryDirectory() as tmp:
        snap = Path(tmp) / snap_name
        for attempt in (1, 2):
            try:
                shutil.copy2(src, snap)
                break
            except PermissionError:
                if attempt == 2:
                    raise
                logging.info("Source file busy, retrying in 2s...")
                time.sleep(2)

        body = (
            f"Snapshot of: {src.name}\n"
            f"Snapshot taken: {ts}\n"
            f"File size: {snap.stat().st_size:,} bytes\n\n"
            f"{summarise(snap)}\n\n"
            f"The running automation continues writing to the original file - "
            f"this attachment is a point-in-time copy."
        )
        send_email(body, snap, snap_name)
        logging.info("Sent %s to %s", snap_name, ", ".join(RECIPIENTS))

    return 0


if __name__ == "__main__":
    sys.exit(main())
