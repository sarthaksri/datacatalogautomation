"""Split the latest comparison report into one workbook per dataset.

For every (dataset_name, output_file) pair in DATASETS, this builds a
workbook that keeps the same sheet tabs, header row, and cell formatting
as the comparison report, but only includes rows whose dataset name in
column A mentions that dataset. Datasets with no matching rows in the
report are skipped (no file written).

Usage:
    python scripts/extract_cpi_iip.py
    python scripts/extract_cpi_iip.py <comparison_report.xlsx>
"""
import sys
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook

_HERE = Path(__file__).resolve().parent

# Sheets in the report we never want in the per-dataset outputs.
_SKIP_SHEETS = {"completed"}

# (match string — matched case-insensitively as a substring of column A,
#  output filename). Trailing years / "NSS Nth Round" suffixes are dropped
#  from the match string so they stay stable across report refreshes.
DATASETS = [
    ("NATIONAL ACCOUNTS STATISTICS NAS",                                  "nas.xlsx"),
    ("PERIODIC LABOUR FORCE SURVEY PLFS",                                 "plfs.xlsx"),
    ("WOMEN AND MEN IN INDIA WMI",                                        "wmi.xlsx"),
    ("ANNUAL SURVEY OF UNINCORPORATED SECTOR ENTERPRISES ASUSE",          "asuse.xlsx"),
    ("ENVISTATS INDIA EI",                                                "envistats_ei.xlsx"),
    ("HOUSEHOLD CONSUMPTION EXPENDITURE SURVEY HCES",                     "hces.xlsx"),
    ("ALL INDIA DEBT AND INVESTMENT SURVEY AIDIS",                        "aidis.xlsx"),
    ("PERSONS WITH DISABILITIES IN INDIA PWDII",                          "pwdii.xlsx"),
    ("ENERGY STATISTICS INDIA ESI",                                       "esi.xlsx"),
    ("SITUATION ASSESSMENT OF AGRICULTURAL HOUSEHOLDS",                   "saahll.xlsx"),
    ("MULTIPLE INDICATOR SURVEY MIS",                                     "mis.xlsx"),
    ("HOUSEHOLD SOCIAL CONSUMPTION ON EDUCATION IN INDIA SCE",            "sce.xlsx"),
    ("SOCIAL CONSUMPTION IN INDIA: HEALTH SCH",                           "sch.xlsx"),
    ("COMPREHENSIVE ANNUAL MODULAR SURVEY CAMS",                          "cams.xlsx"),
    ("DRINKING WATER, SANITATION, HYGIENE AND HOUSING CONDITION IN INDIA DWSHHC", "dwshhc.xlsx"),
    ("TECHNICAL REPORT ON SERVICES SECTOR ENTERPRISES IN INDIA TRSSEI",   "trssei.xlsx"),
    ("COMPREHENSIVE MODULAR SURVEY : EDUCATION CMSE",                     "cmse.xlsx"),
    ("COMPREHENSIVE MODULAR SURVEY TELECOM CMST",                         "cmst.xlsx"),
    ("PILOT STUDY ON CONSTRUCTION ACTIVITIES IN UNINCORPORATED SECTOR ESTABLISHMENTS AND HOUSEHOLDS PSCUEH", "pscueh.xlsx"),
    ("TIME USE SURVEY TUS",                                               "tus.xlsx"),
    ("NUTRITIONAL INTAKE IN INDIA NII",                                   "nii.xlsx"),
    ("ANNUAL SURVEY OF INDUSTRIES ASI",                                   "asi.xlsx"),
    ("SURVEY ON AYUSH SOA",                                               "soa.xlsx"),
    ("CONSUMER PRICE INDEX CPI",                                          "cpi.xlsx"),
    ("INDEX OF INDUSTRIAL PRODUCTION IIP",                                "iip.xlsx"),
]


def _newest_report() -> Path:
    candidates = sorted(_HERE.glob("comparison_report_*.xlsx"))
    if not candidates:
        sys.exit(f"No comparison_report_*.xlsx found in {_HERE}")
    return candidates[-1]


def _resolve_report_path(raw_path: Path) -> Path:
    if raw_path.is_absolute():
        path = raw_path.resolve(strict=False)
    else:
        rel = raw_path
        if raw_path.parts and raw_path.parts[0].lower() == _HERE.name.lower():
            rel = Path(*raw_path.parts[1:])
        path = (_HERE / rel).resolve(strict=False)
    try:
        path.relative_to(_HERE)
    except ValueError:
        sys.exit(f"Report must be inside scripts directory: {_HERE}")
    return path


def _matches(name, dataset_name: str) -> bool:
    return bool(name) and dataset_name in str(name).upper()


def _copy_row_styles(src_ws, dst_ws, src_row: int, dst_row: int, width: int) -> None:
    for col_idx in range(1, width + 1):
        src_cell = src_ws.cell(row=src_row, column=col_idx)
        dst_cell = dst_ws.cell(row=dst_row, column=col_idx)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
    src_dim = src_ws.row_dimensions.get(src_row)
    if src_dim is not None and src_dim.height is not None:
        dst_ws.row_dimensions[dst_row].height = src_dim.height


def _copy_layout(src_ws, dst_ws, width: int) -> None:
    for col_idx in range(1, width + 1):
        letter = dst_ws.cell(row=1, column=col_idx).column_letter
        src_dim = src_ws.column_dimensions.get(letter)
        if src_dim and src_dim.width is not None:
            dst_ws.column_dimensions[letter].width = src_dim.width
    # Don't copy the source's freeze_panes: the report freezes a row near the
    # bottom of its huge sheet (e.g. 'A239108'). On a trimmed sheet that row is
    # past the data, so the entire sheet lands in the frozen pane and won't
    # scroll. Just freeze the header row instead.
    dst_ws.freeze_panes = "A2" if dst_ws.max_row >= 2 else None


def _status_col_idx(src_ws, width: int):
    """Return the 1-based index of the 'Status' header column, or None."""
    for c in range(1, width + 1):
        val = src_ws.cell(row=1, column=c).value
        if val is not None and str(val).strip().lower() == "status":
            return c
    return None


def _is_fail_or_error(value) -> bool:
    if value is None:
        return False
    s = str(value).upper()
    return "FAIL" in s or "ERROR" in s


def _match_dataset(name) -> int:
    """Return the index in DATASETS this row belongs to, or -1 for none.

    First match wins. The match strings are full distinctive dataset names,
    so a row matches at most one entry anyway.
    """
    if not name:
        return -1
    upper = str(name).upper()
    for idx, (dataset_name, _) in enumerate(DATASETS):
        if dataset_name in upper:
            return idx
    return -1


def build(source_path: Path) -> None:
    if not source_path.exists():
        sys.exit(f"File not found: {source_path}")

    print(f"Source report: {source_path.name}\n")
    src_wb = load_workbook(source_path)

    sheets = [s for s in src_wb.sheetnames if s.strip().lower() not in _SKIP_SHEETS]
    widths = {s: src_wb[s].max_column for s in sheets}
    # 1-based index of the "Status" column per sheet (None if the sheet has
    # none — e.g. Mismatches, where every row is already a cell mismatch).
    status_col = {s: _status_col_idx(src_wb[s], widths[s]) for s in sheets}

    # One output workbook per dataset, with a dst worksheet + header for each
    # kept sheet. out_row tracks the next write row per (dataset, sheet).
    out_wbs = []      # parallel to DATASETS: Workbook
    dst_sheets = []   # parallel to DATASETS: {sheet_name: worksheet}
    out_rows = []     # parallel to DATASETS: {sheet_name: next_row_int}
    kept_counts = []  # parallel to DATASETS: {sheet_name: kept_int}
    fe_counts = []    # parallel to DATASETS: {sheet_name: fail_or_error_int}

    for _ in DATASETS:
        wb = Workbook()
        wb.remove(wb.active)
        ds, rows, kept, fe = {}, {}, {}, {}
        for s in sheets:
            src_ws = src_wb[s]
            dst_ws = wb.create_sheet(s)
            width = widths[s]
            if src_ws.max_row >= 1:
                for c in range(1, width + 1):
                    dst_ws.cell(row=1, column=c, value=src_ws.cell(row=1, column=c).value)
                _copy_row_styles(src_ws, dst_ws, 1, 1, width)
            ds[s], rows[s], kept[s], fe[s] = dst_ws, 2, 0, 0
        out_wbs.append(wb)
        dst_sheets.append(ds)
        out_rows.append(rows)
        kept_counts.append(kept)
        fe_counts.append(fe)

    # SINGLE PASS per sheet: route each source row to its dataset's workbook.
    # Iterate cells (not values_only) so each row's fills/fonts/borders — the
    # FAIL/ERROR/PASS colour coding — carry over too.
    for s in sheets:
        src_ws = src_wb[s]
        width = widths[s]
        sc = status_col[s]
        for row in src_ws.iter_rows(min_row=2):
            idx = _match_dataset(row[0].value)
            if idx < 0:
                continue
            dst_ws = dst_sheets[idx][s]
            out_row = out_rows[idx][s]
            for col_idx, src_cell in enumerate(row, start=1):
                dst_ws.cell(row=out_row, column=col_idx, value=src_cell.value)
            _copy_row_styles(src_ws, dst_ws, row[0].row, out_row, width)
            out_rows[idx][s] = out_row + 1
            kept_counts[idx][s] += 1
            if sc is not None and _is_fail_or_error(row[sc - 1].value):
                fe_counts[idx][s] += 1

    # Finalise layout + save (skip datasets with no matched rows).
    written = skipped = grand_total = 0
    for idx, (dataset_name, out_file) in enumerate(DATASETS):
        total = sum(kept_counts[idx].values())
        if total == 0:
            print(f"SKIP: {out_file:18s} (no rows for {dataset_name!r})")
            skipped += 1
            continue
        wb = out_wbs[idx]
        for s in sheets:
            _copy_layout(src_wb[s], dst_sheets[idx][s], widths[s])
        out_path = _HERE / out_file
        try:
            wb.save(out_path)
        except PermissionError:
            sys.exit(f"Cannot write {out_path}. Close the file in Excel and run again.")
        # Per-sheet line: total rows + fail/error count (or n/a where the
        # sheet carries no Status column).
        print(f"OK: {out_file}  ({total} rows total)")
        for s in sheets:
            rows_s = kept_counts[idx][s]
            fe_s = f"{fe_counts[idx][s]} fail/error" if status_col[s] is not None else "fail/error n/a"
            print(f"       {s:12s}: {rows_s:6d} rows, {fe_s}")
        written += 1
        grand_total += total

    print(f"\nDone: wrote {written} workbook(s), skipped {skipped}, "
          f"{grand_total} total rows.")


if __name__ == "__main__":
    report_path = _resolve_report_path(Path(sys.argv[1])) if len(sys.argv) > 1 else _newest_report()
    build(report_path)
