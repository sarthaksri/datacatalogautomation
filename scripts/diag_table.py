"""Diagnostic: for a given dataset, drive the live site through the first
few tables and dump exactly what ExcelHandler and TableViewer each produce,
plus the raw download, so we can see WHY comparisons fail.

Run from the project root with the venv python:
    venv\\Scripts\\python.exe scripts\\diag_table.py "PERIODIC LABOUR FORCE SURVEY PLFS" 2
    venv\\Scripts\\python.exe scripts\\diag_table.py "TIME USE SURVEY TUS" 1

Writes a report to scripts/diag_<acronym>.txt and saves raw .xlsx downloads
to scripts/diag_downloads/.
"""
import sys
from pathlib import Path

# Make project modules importable when run from scripts/
_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

import config
from modules.browser import BrowserManager
from modules.catalogue_page import CataloguePage
from modules.dataset_page import DatasetPage
from modules.excel_handler import ExcelHandler
from modules.table_viewer import TableViewer
from modules.comparator import compare
from modules.header_detect import pick_header_idx

import pandas as pd

_HERE = Path(__file__).resolve().parent
_DL = _HERE / "diag_downloads"
_DL.mkdir(exist_ok=True)


def _fmt_rows(rows, n=6, w=10):
    out = []
    for r in rows[:n]:
        cells = [(str(c)[:w]) for c in r[:12]]
        out.append("   | " + " | ".join(cells))
    return "\n".join(out)


def _raw_excel_dump(path, lines):
    """Show the raw rows + what openpyxl/pandas see, before header picking."""
    lines.append(f"  -- RAW EXCEL ({path.name}) --")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        lines.append(f"  openpyxl sheets: {wb.sheetnames}")
    except Exception as e:
        lines.append(f"  openpyxl FAILED: {e!r}")
    try:
        df = pd.read_excel(path, sheet_name=0, header=None, dtype=str).fillna("")
        rows = [list(r) for r in df.values.tolist()]
        lines.append(f"  pandas shape: {df.shape}; picked header idx: {pick_header_idx(rows)}")
        lines.append("  first 8 raw rows:")
        lines.append(_fmt_rows(rows, 8))
    except Exception as e:
        lines.append(f"  pandas read FAILED: {e!r}")


def diag(dataset_substr: str, max_tables: int):
    acro = dataset_substr.strip().split()[-1].lower()
    out_path = _HERE / f"diag_{acro}.txt"
    lines = [f"DATASET FILTER: {dataset_substr!r}", "=" * 70]

    with BrowserManager(headless=True, download_dir=_DL, slow_mo=0) as bm:
        bm.page.set_default_timeout(60_000)
        catalogue = CataloguePage(
            page=bm.page, url=config.CATALOGUE_URL,
            card_selectors=config.CATALOGUE_CARD_SELECTORS,
        )
        catalogue.load()
        datasets = catalogue.get_datasets()
        needle = dataset_substr.lower()
        match = next((d for d in datasets if needle in d["name"].lower()), None)
        if not match:
            lines.append(f"NO MATCH. Available: {[d['name'] for d in datasets]}")
            out_path.write_text("\n".join(lines), encoding="utf-8")
            return
        lines.append(f"Matched dataset: {match['name']!r} (index {match['index']})")
        catalogue.click_dataset_by_index(match["index"])

        dp = DatasetPage(page=bm.page, download_dir=_DL)
        dp.wait_for_table()
        rows = dp.get_rows()
        dataset_url = bm.page.url

        for i, row_meta in enumerate(rows[:max_tables]):
            tno = row_meta.get("table_no", str(i))
            lines.append("\n" + "#" * 70)
            lines.append(f"TABLE {i}: [{tno}] {row_meta.get('table_name','')[:80]}")
            lines.append("#" * 70)

            # Inspect the download anchor + try a direct HTTP GET of its href.
            try:
                card = bm.page.locator(dp.CARD_SEL).nth(i)
                a = card.locator(dp.DOWNLOAD_SEL).first
                href = a.get_attribute("href")
                lines.append(f"  download href: {href!r}")
                if href:
                    full = href if href.startswith("http") else config.BASE_URL + href
                    # Fetch through the browser's own network stack (handles the
                    # MoSPI SSL legacy-renegotiation that APIRequestContext trips on).
                    js = """
                    async (u) => {
                        try {
                            const r = await fetch(u);
                            const b = await r.arrayBuffer();
                            const bytes = new Uint8Array(b).slice(0,4);
                            return {status:r.status, ct:r.headers.get('content-type'),
                                    len:b.byteLength, magic:Array.from(bytes)};
                        } catch(e) { return {error:String(e)}; }
                    }"""
                    for label, u in [("raw", full),
                                     ("encoded", config.BASE_URL + "/" + "/".join(
                                         __import__("urllib.parse", fromlist=["quote"]).quote(p)
                                         for p in href.lstrip("/").split("/")))]:
                        res = bm.page.evaluate(js, u)
                        lines.append(f"  browser fetch [{label}] {u[:95]} -> {res}")
            except Exception as e:
                lines.append(f"  href/direct-GET inspect failed: {e!r}")

            excel_path = dp.download_excel(i, row_meta.get("table_name", ""), tno)

            dp.click_view_table(i)
            tv = TableViewer(page=bm.page, debug_dump_dir=None)
            web = tv.get_all_data()

            # back to listing
            if "/tableview/" in bm.page.url or bm.page.url != dataset_url:
                bm.page.goto(dataset_url, wait_until="networkidle")
            dp.wait_for_table()

            if excel_path:
                _raw_excel_dump(excel_path, lines)
            else:
                lines.append("  -- NO EXCEL DOWNLOADED --")

            excel = ExcelHandler.read(excel_path) if excel_path else None
            lines.append("\n  -- PARSED EXCEL --")
            if excel:
                lines.append(f"  header_row_idx={excel.get('header_row_idx')}  "
                             f"cols={len(excel['headers'])}  rows={len(excel['rows'])}")
                lines.append(f"  headers: {excel['headers'][:12]}")
                lines.append(_fmt_rows(excel["rows"]))
            else:
                lines.append("  Excel read returned None (THIS is the 'could not be read' error)")

            lines.append("\n  -- PARSED WEB --")
            lines.append(f"  cols={len(web['headers'])}  rows={len(web['rows'])}")
            lines.append(f"  headers: {web['headers'][:12]}")
            lines.append(_fmt_rows(web["rows"]))

            res = compare(excel, web)
            lines.append(f"\n  -- COMPARE --  status={res.status} mism={res.mismatch_count}")
            if res.error:
                lines.append(f"  error: {res.error}")
            if res.missing_cols:
                lines.append(f"  missing_cols (Excel only): {res.missing_cols[:12]}")
            if res.extra_cols:
                lines.append(f"  extra_cols (Web only): {res.extra_cols[:12]}")
            lines.append(f"  excel_extra_rows={res.excel_extra_rows} web_extra_rows={res.web_extra_rows}")
            for m in res.mismatches[:20]:
                lines.append(f"    mism row{m.row} [{m.row_label[:30]}] col={m.column!r}: "
                             f"Excel={m.excel_value!r} Web={m.web_value!r}")
            if excel and excel["rows"]:
                lines.append("  EXCEL tail rows:")
                lines.append(_fmt_rows(excel["rows"][-4:]))
            if web and web["rows"]:
                lines.append("  WEB tail rows:")
                lines.append(_fmt_rows(web["rows"][-4:]))

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    substr = sys.argv[1] if len(sys.argv) > 1 else "PERIODIC LABOUR FORCE SURVEY PLFS"
    n = int(sys.argv[2]) if len(sys.argv) > 2 else 1
    diag(substr, n)
