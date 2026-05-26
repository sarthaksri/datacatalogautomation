import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .comparator import ComparisonResult, Mismatch

log = logging.getLogger(__name__)

# Palette
_BLUE       = "FF4472C4"
_WHITE      = "FFFFFFFF"
_LIGHT_GRAY = "FFF2F2F2"
_LIGHT_GREEN = "FFC6EFCE"
_LIGHT_RED   = "FFFFC7CE"
_LIGHT_AMBER = "FFFFEB9C"
_DARK_RED    = "FF9C0006"
_DARK_GREEN  = "FF276221"
_DARK_AMBER  = "FF9C5700"


def _hdr(cell, bg=_BLUE) -> None:
    cell.font      = Font(bold=True, color=_WHITE, size=11)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width  = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 70)


class Reporter:
    """
    Accumulates per-table results and writes a checkpointable .xlsx report.

    Pass `report_path` at construction time to enable incremental saves and
    resume-from-disk. If the file already exists, prior entries and the set
    of fully-completed datasets are loaded so a follow-up run can pick up
    where the previous one left off.
    """

    def __init__(self, report_path: Optional[Path] = None) -> None:
        self._entries: List[Dict] = []
        self._completed_datasets: Set[str] = set()
        self.report_path: Optional[Path] = report_path
        if report_path and report_path.exists():
            try:
                self._load_existing(report_path)
                log.info(
                    "Loaded %d prior entries and %d completed dataset(s) from %s",
                    len(self._entries), len(self._completed_datasets), report_path,
                )
            except Exception as exc:
                log.warning("Could not load %s — starting fresh: %s", report_path, exc)
                self._entries = []
                self._completed_datasets = set()

    # ── Public API ────────────────────────────────────────────────────────────

    def add_result(self, dataset: str, meta: Dict, result: ComparisonResult) -> None:
        self._entries.append({"dataset": dataset, "meta": meta, "result": result})

    def add_error(self, dataset: str, label: str, error_msg: str) -> None:
        r = ComparisonResult(status="ERROR", error=error_msg)
        empty_meta = {
            "table_name": label, "table_no": "",
            "product": "", "category": "", "release_date": "",
        }
        self._entries.append({"dataset": dataset, "meta": empty_meta, "result": r})

    def mark_dataset_complete(self, dataset_name: str) -> None:
        """Flag a dataset as fully processed (all pages iterated)."""
        self._completed_datasets.add(dataset_name)

    def completed_datasets(self) -> Set[str]:
        return set(self._completed_datasets)

    def completed_tables(self) -> Set[Tuple[str, str]]:
        """(dataset_name, table_no) pairs already in the report."""
        return {(e["dataset"], e["meta"].get("table_no", "")) for e in self._entries}

    def save(self) -> Optional[Path]:
        """Write the current state to `self.report_path` (overwriting)."""
        if not self.report_path:
            return None
        self.report_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        self._write_summary(wb)
        self._write_mismatches(wb)
        self._write_completed(wb)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        try:
            wb.save(self.report_path)
        except PermissionError:
            log.warning(
                "Could not save %s (file open in Excel?); will retry on next save",
                self.report_path,
            )
            return None
        return self.report_path

    def generate(self, report_dir: Path = Path("reports")) -> Path:
        """Back-compat: pick a path if not set, then save."""
        if not self.report_path:
            report_dir.mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.report_path = report_dir / f"comparison_report_{ts}.xlsx"
        path = self.save()
        if path:
            log.info("Report saved: %s", path)
            print(f"\n✓ Report saved → {path}")
        return self.report_path

    # ── Reload from disk ──────────────────────────────────────────────────────

    def _load_existing(self, path: Path) -> None:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        # Summary → reconstruct entry stubs keyed by (dataset, table_no).
        idx: Dict[Tuple[str, str], Dict] = {}
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                (dataset, table_no, table_name, product, category,
                 release_date, status, total_rows, _mm_count, notes) = (
                    list(row[:10]) + [None] * (10 - len(row[:10]))
                )
                meta = {
                    "table_no":     str(table_no or ""),
                    "table_name":   str(table_name or ""),
                    "product":      str(product or ""),
                    "category":     str(category or ""),
                    "release_date": str(release_date or ""),
                }
                result = ComparisonResult(
                    status=str(status or "PASS"),
                    total_rows=int(total_rows or 0),
                )
                entry = {
                    "dataset": str(dataset),
                    "meta":    meta,
                    "result":  result,
                    "_raw_notes": notes or "",   # preserve notes verbatim
                }
                self._entries.append(entry)
                idx[(entry["dataset"], meta["table_no"])] = entry

        # Mismatches → attach to their parent entry.
        if "Mismatches" in wb.sheetnames:
            ws = wb["Mismatches"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                (dataset, table_no, _table_name, excel_cell, row_num,
                 row_label, column, excel_value, web_value) = (
                    list(row[:9]) + [None] * (9 - len(row[:9]))
                )
                key = (str(dataset), str(table_no or ""))
                entry = idx.get(key)
                if entry is None:
                    continue
                try:
                    row_int = int(row_num or 0)
                except (TypeError, ValueError):
                    row_int = 0
                entry["result"].mismatches.append(Mismatch(
                    row=row_int,
                    column=str(column or ""),
                    excel_value=str(excel_value or ""),
                    web_value=str(web_value or ""),
                    excel_cell=str(excel_cell or ""),
                    row_label=str(row_label or ""),
                ))

        # Completed datasets sheet.
        if "Completed" in wb.sheetnames:
            ws = wb["Completed"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    self._completed_datasets.add(str(row[0]))

        wb.close()

    # ── Sheet builders ────────────────────────────────────────────────────────

    def _write_summary(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Summary", 0)
        ws.freeze_panes = "A2"

        cols = [
            "Dataset", "Table No.", "Table Name", "Product",
            "Category", "Release Date", "Status",
            "Total Rows", "Mismatches", "Notes",
        ]
        for ci, h in enumerate(cols, 1):
            _hdr(ws.cell(row=1, column=ci, value=h))

        for ri, entry in enumerate(self._entries, 2):
            meta: Dict            = entry["meta"]
            res: ComparisonResult = entry["result"]

            ws.cell(ri, 1,  entry["dataset"])
            ws.cell(ri, 2,  meta.get("table_no",      ""))
            ws.cell(ri, 3,  meta.get("table_name",    ""))
            ws.cell(ri, 4,  meta.get("product",       ""))
            ws.cell(ri, 5,  meta.get("category",      ""))
            ws.cell(ri, 6,  meta.get("release_date",  ""))

            sc = ws.cell(ri, 7, res.status)
            if res.status == "PASS":
                sc.fill = PatternFill("solid", fgColor=_LIGHT_GREEN)
                sc.font = Font(bold=True, color=_DARK_GREEN)
            elif res.status == "FAIL":
                sc.fill = PatternFill("solid", fgColor=_LIGHT_RED)
                sc.font = Font(bold=True, color=_DARK_RED)
            else:
                sc.fill = PatternFill("solid", fgColor=_LIGHT_AMBER)
                sc.font = Font(bold=True, color=_DARK_AMBER)
            sc.alignment = Alignment(horizontal="center")

            ws.cell(ri, 8,  res.total_rows)
            ws.cell(ri, 9,  res.mismatch_count)

            # Reloaded entries keep their original notes verbatim; freshly
            # computed entries derive notes from the live ComparisonResult.
            raw_notes = entry.get("_raw_notes")
            if raw_notes:
                ws.cell(ri, 10, str(raw_notes))
            else:
                notes: List[str] = []
                if res.error:
                    notes.append(f"Error: {res.error}")
                if res.col_mismatch:
                    if res.missing_cols:
                        notes.append(f"Missing cols (Excel only): {', '.join(res.missing_cols)}")
                    if res.extra_cols:
                        notes.append(f"Extra cols (Web only): {', '.join(res.extra_cols)}")
                if res.excel_extra_rows:
                    notes.append(f"Excel has {res.excel_extra_rows} extra row(s)")
                if res.web_extra_rows:
                    notes.append(f"Web has {res.web_extra_rows} extra row(s)")
                ws.cell(ri, 10, "; ".join(notes))

            if ri % 2 == 0:
                for ci in [1, 2, 3, 4, 5, 6, 8, 9, 10]:
                    ws.cell(ri, ci).fill = PatternFill("solid", fgColor=_LIGHT_GRAY)

        _auto_width(ws)
        ws.row_dimensions[1].height = 30

    def _write_mismatches(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Mismatches", 1)
        ws.freeze_panes = "A2"

        cols = [
            "Dataset", "Table No.", "Table Name",
            "Excel Cell", "Row #", "Row Label", "Column",
            "Excel Value", "Web Value",
        ]
        for ci, h in enumerate(cols, 1):
            _hdr(ws.cell(row=1, column=ci, value=h))

        rn = 2
        for entry in self._entries:
            meta: Dict            = entry["meta"]
            res: ComparisonResult = entry["result"]

            for mm in res.mismatches:
                ws.cell(rn, 1, entry["dataset"])
                ws.cell(rn, 2, meta.get("table_no",   ""))
                ws.cell(rn, 3, meta.get("table_name", ""))
                ws.cell(rn, 4, mm.excel_cell)
                ws.cell(rn, 5, mm.row)
                ws.cell(rn, 6, mm.row_label)
                ws.cell(rn, 7, mm.column)

                ec = ws.cell(rn, 8, mm.excel_value)
                ec.fill = PatternFill("solid", fgColor=_LIGHT_RED)
                ec.font = Font(color=_DARK_RED)

                wc = ws.cell(rn, 9, mm.web_value)
                wc.fill = PatternFill("solid", fgColor=_LIGHT_AMBER)
                wc.font = Font(color=_DARK_AMBER)

                if rn % 2 == 0:
                    for ci in [1, 2, 3, 4, 5, 6, 7]:
                        ws.cell(rn, ci).fill = PatternFill("solid", fgColor=_LIGHT_GRAY)
                rn += 1

        if rn == 2:
            ws.cell(2, 1, "No mismatches found — all data matches.")

        _auto_width(ws)
        ws.row_dimensions[1].height = 30

    def _write_completed(self, wb: openpyxl.Workbook) -> None:
        """
        Sheet listing datasets whose every page has been iterated end-to-end.
        On `--continue`, datasets in this list are skipped wholesale by main.
        """
        ws = wb.create_sheet("Completed")
        _hdr(ws.cell(row=1, column=1, value="Dataset"))
        for ri, name in enumerate(sorted(self._completed_datasets), 2):
            ws.cell(ri, 1, name)
        _auto_width(ws)
        ws.row_dimensions[1].height = 24
